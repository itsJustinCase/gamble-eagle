#!/usr/bin/env python3
"""
Brazil Licensed Gambling Sites Scraper
=======================================
Source: Secretaria de Prêmios e Apostas (SPA) — Ministério da Fazenda
        https://www.gov.br/fazenda/pt-br/composicao/orgaos/secretaria-de-premios-e-apostas/lista-de-empresas/
"""

import os
import re
import csv
import io
import requests
from datetime import datetime
from bs4 import BeautifulSoup
import openpyxl

try:
    from zoneinfo import ZoneInfo
    _PARIS = ZoneInfo('Europe/Paris')
    def _paris_now():
        return datetime.now(_PARIS)
except ImportError:
    import pytz
    _PARIS = pytz.timezone('Europe/Paris')
    def _paris_now():
        return datetime.now(_PARIS)

# ── Config ────────────────────────────────────────────────────────────────────

TARGET_URLS = [
    "https://www.gov.br/fazenda/pt-br/composicao/orgaos/secretaria-de-premios-e-apostas/lista-de-empresas/empresas-autorizadas",
    "https://www.gov.br/fazenda/pt-br/composicao/orgaos/secretaria-de-premios-e-apostas/lista-de-empresas/autorizadas-por-determinacao-judicial",
]

BASE_URL = "https://www.gov.br"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "pt-BR,pt;q=0.9,en;q=0.8",
}

MIN_EXPECTED = 20

# ── Canonical CSV writer ──────────────────────────────────────────────────────

def write_canonical_csv(urls, filepath):
    stamp = _paris_now().strftime('%Y%m%d %H:%M')
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        f.write(stamp + '\n')
        for url in urls:
            f.write(url.strip() + '\n')
    print(f"💾  Saved {len(urls)} URLs → {filepath}  (stamp: {stamp})")

# ── Domain splitting & cleaning ───────────────────────────────────────────────

def clean_domain(raw):
    d = raw.strip().lower()
    d = re.sub(r'^https?://', '', d)
    d = re.sub(r'^www\.', '', d)
    d = d.rstrip('/')
    return d

def is_valid(domain):
    if not domain:
        return False
    low = domain.lower().strip()
    if low in ('a definir', 'a definir.', '', '-', 'n/a', 'nd'):
        return False
    if '.' not in low:
        return False
    if len(low) < 4:
        return False
    return True

def process_raw_text(text):
    """Splits grouped domain strings (by whitespace, linebreaks, commas, semicolons, slashes)
    into distinct domain entries."""
    if not text:
        return []
    
    extracted = []
    # Split by newlines, spaces, commas, semicolons, slashes, or bullet separators
    tokens = re.split(r'[\r\n,;\s/]+', str(text))
    for token in tokens:
        item = token.strip()
        if not item or 'a definir' in item.lower():
            continue
        cleaned = clean_domain(item)
        if is_valid(cleaned):
            extracted.append(cleaned)
    return extracted

# ── Parsers ───────────────────────────────────────────────────────────────────

def extract_from_html_tables(soup):
    domains = []
    tables = soup.find_all('table')

    for table in tables:
        rows = table.find_all('tr')
        if not rows:
            continue

        domain_col = None
        header_row_idx = None

        # Locate the header row containing 'domínio' or 'domínios'
        for i, row in enumerate(rows):
            cells = [cell.get_text(strip=True) for cell in row.find_all(['th', 'td'])]
            for j, text in enumerate(cells):
                low_text = text.lower()
                if 'dom' in low_text and ('nio' in low_text or 'nios' in low_text):
                    domain_col = j
                    header_row_idx = i
                    break
            if domain_col is not None:
                break

        if domain_col is None:
            continue

        # Extract values from rows following the header
        for row in rows[header_row_idx + 1:]:
            cells = row.find_all(['td', 'th'])
            if len(cells) <= domain_col:
                continue

            # Get raw cell text retaining line breaks
            raw_cell = cells[domain_col].get_text(separator='\n', strip=True)
            domains.extend(process_raw_text(raw_cell))

    return domains

def extract_from_excel(content):
    domains = []
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        domain_col = None
        header_row_idx = None

        rows = list(sheet.iter_rows(values_only=True))
        for i, row in enumerate(rows):
            for j, cell in enumerate(row):
                if cell and isinstance(cell, str):
                    low_cell = cell.lower()
                    if 'dom' in low_cell and ('nio' in low_cell or 'nios' in low_cell):
                        domain_col = j
                        header_row_idx = i
                        break
            if domain_col is not None:
                break

        if domain_col is None:
            continue

        for row in rows[header_row_idx + 1:]:
            if len(row) <= domain_col or row[domain_col] is None:
                continue
            domains.extend(process_raw_text(row[domain_col]))

    return domains

def extract_from_csv(content_bytes):
    content = content_bytes.decode('utf-8-sig', errors='replace')
    domains = []
    reader = csv.reader(io.StringIO(content), delimiter=';')
    rows = list(reader)

    domain_col = None
    header_row_idx = None
    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            low_cell = cell.lower()
            if 'dom' in low_cell and ('nio' in low_cell or 'nios' in low_cell):
                domain_col = j
                header_row_idx = i
                break
        if domain_col is not None:
            break

    if domain_col is None:
        return []

    for row in rows[header_row_idx + 1:]:
        if len(row) <= domain_col:
            continue
        domains.extend(process_raw_text(row[domain_col]))

    return domains

def process_page(session, url):
    print(f"🌐  Fetching page: {url}")
    r = session.get(url, headers=HEADERS, timeout=30)
    r.raise_for_status()

    soup = BeautifulSoup(r.content, 'html.parser')
    domains = []

    # 1. Primary: Extract directly from inline HTML tables
    html_domains = extract_from_html_tables(soup)
    domains.extend(html_domains)

    # 2. Fallback: Search for downloadable spreadsheets linked on the page
    patterns = [
        re.compile(r'\.(xlsx|csv)$', re.IGNORECASE),
        re.compile(r'planilha', re.IGNORECASE)
    ]
    
    file_urls = []
    for a in soup.find_all('a', href=True):
        href = a['href']
        text = a.get_text()
        if any(p.search(href) or p.search(text) for p in patterns):
            file_url = href if href.startswith('http') else BASE_URL + href
            if file_url not in file_urls:
                file_urls.append(file_url)

    for file_url in file_urls:
        try:
            print(f"    📥  Found downloadable file: {file_url.split('/')[-1]}")
            fr = session.get(file_url, headers=HEADERS, timeout=30)
            fr.raise_for_status()
            if file_url.lower().endswith('.xlsx'):
                domains.extend(extract_from_excel(fr.content))
            else:
                domains.extend(extract_from_csv(fr.content))
        except Exception as e:
            print(f"    ❌  Failed to process file {file_url}: {e}")

    print(f"    📊  Extracted {len(domains)} domains from {url.split('/')[-1]}")
    return domains

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("🇧🇷  BRAZIL LICENSED GAMBLING SITES SCRAPER (SPA)")
    print("=" * 60)

    session = requests.Session()
    all_domains = []

    for url in TARGET_URLS:
        try:
            domains = process_page(session, url)
            all_domains.extend(domains)
        except Exception as e:
            print(f"❌  Failed to process page {url}: {e}")

    seen = set()
    unique = []
    for d in all_domains:
        if d not in seen:
            seen.add(d)
            unique.append(d)
    unique.sort()

    print(f"\n📊  Total unique domains combined: {len(unique)}")

    if len(unique) < MIN_EXPECTED:
        print(f"❌  Only {len(unique)} domains found — below minimum threshold ({MIN_EXPECTED}).")
        return

    print(f"\n🔍  First 10 domains:")
    for d in unique[:10]:
        print(f"    {d}")
    if len(unique) > 10:
        print(f"    ... and {len(unique) - 10} more")

    write_canonical_csv(unique, 'brazil.csv')
    print("✅  Done.")

if __name__ == "__main__":
    main()
    if not os.environ.get("CI"):
        input("\nPress Enter to close...")
